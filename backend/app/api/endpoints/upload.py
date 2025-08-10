from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from typing import List
import uuid
import os
from pathlib import Path
import shutil
import re
import logging
from datetime import datetime

from app.models.comparison import SessionResponse
from app.core.config import settings
from app.services.dual_parser import DualExcelParser
from app.services.universal_parser import UniversalExcelParser
import openpyxl

router = APIRouter()

# Set up logger
logger = logging.getLogger(__name__)

# In-memory session storage (will be replaced with proper database)
active_sessions = {}

@router.post("/upload-models", response_model=SessionResponse)
async def upload_model_pair(
    old_file: UploadFile = File(..., description="Old model Excel file"),
    new_file: UploadFile = File(..., description="New model Excel file")
):
    """
    Upload a pair of Excel financial models for comparison
    """
    # Validate file types
    if not _is_valid_file(old_file.filename) or not _is_valid_file(new_file.filename):
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Only .xlsx and .xls files are supported."
        )
    
    # Create session
    session_id = str(uuid.uuid4())
    
    try:
        # Create upload directory
        upload_dir = Path(settings.UPLOAD_DIR) / session_id
        upload_dir.mkdir(parents=True, exist_ok=True)
        
        # Save files
        old_file_path = upload_dir / f"old_{old_file.filename}"
        new_file_path = upload_dir / f"new_{new_file.filename}"
        
        await _save_upload_file(old_file, old_file_path)
        await _save_upload_file(new_file, new_file_path)
        
        # Store session info
        active_sessions[session_id] = {
            "old_file_path": old_file_path,
            "new_file_path": new_file_path,
            "old_filename": old_file.filename,
            "new_filename": new_file.filename,
            "status": "processing",  # Start processing immediately
            "created_at": datetime.now().isoformat()
        }
        
        logger.info(f"🔄 Session {session_id}: uploaded → processing")
        
        # Start processing immediately
        try:
            # Initialize parser
            parser = DualExcelParser()
            
            logger.info(f"📁 Processing files: {old_file.filename} vs {new_file.filename}")
            
            # Parse models with actual implementation
            old_model, new_model, consistency_check = await parser.parse_model_pair(
                Path(old_file_path),
                Path(new_file_path)
            )
            
            # Store parsed models
            active_sessions[session_id]["old_model"] = old_model.dict()  # Convert to dict for JSON serialization
            active_sessions[session_id]["new_model"] = new_model.dict()
            active_sessions[session_id]["consistency_check"] = consistency_check.dict()
            active_sessions[session_id]["status"] = "completed"
            
            logger.info(f"✅ Session {session_id}: processing → completed")
            logger.info(f"   Compatibility score: {consistency_check.compatibility_score}")
            logger.info(f"   Structure match: {consistency_check.structure_match}")
            
            return SessionResponse(
                session_id=session_id,
                status="completed",
                message="Files processed successfully. Ready for analysis.",
                processing_time_estimate=0
            )
            
        except Exception as e:
            active_sessions[session_id]["status"] = "failed"
            active_sessions[session_id]["error"] = str(e)
            
            logger.error(f"💥 Session {session_id}: processing → failed: {str(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            
            return SessionResponse(
                session_id=session_id,
                status="failed",
                message=f"Processing failed: {str(e)}",
                processing_time_estimate=0
            )
        
    except Exception as e:
        # Clean up on error
        if session_id in active_sessions:
            del active_sessions[session_id]
        
        raise HTTPException(
            status_code=500,
            detail=f"Failed to upload files: {str(e)}"
        )

@router.post("/process-models/{session_id}")
async def process_models(session_id: str):
    """
    Start processing the uploaded model pair
    """
    if session_id not in active_sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = active_sessions[session_id]
    
    try:
        # Update status to processing
        session["status"] = "processing"
        
        # Initialize parser
        parser = DualExcelParser()
        
        # Parse models with actual implementation
        old_model, new_model, consistency_check = await parser.parse_model_pair(
            session["old_file_path"],
            session["new_file_path"]
        )
        
        # Store parsed models
        session["old_model"] = old_model.dict()  # Convert to dict for JSON serialization
        session["new_model"] = new_model.dict()
        session["consistency_check"] = consistency_check.dict()
        session["status"] = "completed"
        
        return {
            "message": "Processing completed successfully", 
            "session_id": session_id,
            "compatibility_score": consistency_check.compatibility_score,
            "structure_match": consistency_check.structure_match
        }
        
    except Exception as e:
        session["status"] = "error"
        session["error"] = str(e)
        
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process models: {str(e)}"
        )

@router.get("/session/{session_id}/status")
async def get_session_status(session_id: str):
    """
    Get the current status of a processing session
    """
    if session_id not in active_sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = active_sessions[session_id]
    return {
        "session_id": session_id,
        "status": session["status"],
        "old_filename": session["old_filename"],
        "new_filename": session["new_filename"],
        "created_at": session.get("created_at")
    }

@router.get("/session/{session_id}/sheets")
async def get_available_sheets(session_id: str):
    """
    Get list of sheet names from both uploaded Excel files
    """
    if session_id not in active_sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = active_sessions[session_id]
    
    try:
        old_file_path = session.get("old_file_path")
        new_file_path = session.get("new_file_path")
        
        if not old_file_path or not new_file_path:
            raise HTTPException(status_code=400, detail="File paths not found in session")
        
        # Get sheet names from both files
        old_sheets = _get_sheet_names(old_file_path)
        new_sheets = _get_sheet_names(new_file_path)
        
        return {
            "session_id": session_id,
            "old_model_sheets": old_sheets,
            "new_model_sheets": new_sheets,
            "common_sheets": list(set(old_sheets) & set(new_sheets))
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to read sheet names: {str(e)}"
        )

def _get_sheet_names(file_path):
    """Extract sheet names from Excel file"""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names
    except Exception as e:
        raise Exception(f"Could not read sheets from {file_path}: {str(e)}")

@router.post("/session/{session_id}/select-sheets")
async def select_financial_statement_sheets(
    session_id: str, 
    sheet_selection: dict  # {"income_statement": "IS Sheet", "balance_sheet": "BS Sheet", "cash_flow": "CF Sheet"}
):
    """
    Select which sheets contain financial statements and detect periods
    """
    if session_id not in active_sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = active_sessions[session_id]
    
    try:
        # Store sheet selections
        session["selected_sheets"] = sheet_selection
        
        # Use universal parser to get periods from selected sheets
        parser = UniversalExcelParser()
        old_file_path = session.get("old_file_path")
        
        if not old_file_path:
            raise HTTPException(status_code=400, detail="No uploaded files found")
        
        # Parse just to get periods (from old model as baseline)
        old_statements = parser.parse_financial_statements(old_file_path, sheet_selection)
        
        # Collect all periods from all selected statements
        all_periods = set()
        period_details = {}
        
        for statement_type, statement in old_statements.items():
            all_periods.update(statement.periods)
            period_details[statement_type] = {
                "periods_found": len(statement.periods),
                "periods": statement.periods
            }
        
        periods_list = sorted(list(all_periods))
        
        # Categorize periods for user review
        annual_periods = []
        quarterly_periods = []
        other_periods = []
        
        for period in periods_list:
            if any(q in period.upper() for q in ['1Q', '2Q', '3Q', '4Q']) or 'Q' in period.upper():
                quarterly_periods.append(period)
            elif any(pattern in period for pattern in ['FY', '19', '20', '21', '22', '23', '24', '25', '26', '27']):
                annual_periods.append(period)
            else:
                other_periods.append(period)
        
        # Store detailed period analysis for review
        period_analysis = {
            "total_periods": len(periods_list),
            "annual_periods": annual_periods,
            "quarterly_periods": quarterly_periods,
            "other_periods": other_periods,
            "by_statement": period_details,
            "suggested_templates": []
        }
        
        # Generate template suggestions based on detected patterns (NO hardcoded bank logic)
        if quarterly_periods:
            # Analyze actual patterns found in the data
            fy_quarterly_pattern = any('FY' in p and 'Q' in p for p in quarterly_periods[:5])
            simple_quarterly_pattern = any(re.match(r'^\d+Q\d{2}', p) for p in quarterly_periods[:5])
            q_format_pattern = any(re.match(r'^Q\d', p) for p in quarterly_periods[:5])
            
            if fy_quarterly_pattern:
                period_analysis["suggested_templates"].append({
                    "name": "Fiscal Year Quarterly",
                    "pattern": "FY{Q}Q{YY}[E]",
                    "example": "FY1Q25, FY2Q25E",
                    "description": "Fiscal year quarterly format with optional estimate suffix"
                })
            
            if simple_quarterly_pattern:
                period_analysis["suggested_templates"].append({
                    "name": "Standard Quarterly", 
                    "pattern": "{Q}Q{YY}[E]",
                    "example": "1Q25, 2Q25E",
                    "description": "Standard quarterly format with optional estimate suffix"
                })
                
            if q_format_pattern:
                period_analysis["suggested_templates"].append({
                    "name": "Quarter Year Format",
                    "pattern": "Q{Q} {YYYY}",
                    "example": "Q1 2025, Q2 2025",
                    "description": "Quarter followed by full year"
                })
        
        session["period_analysis"] = period_analysis
        session["periods_need_review"] = len(periods_list) < 50  # Threshold for review
        
        logger.info(f"Sheet selection completed. Periods found: {len(periods_list)} ({len(quarterly_periods)} quarterly, {len(annual_periods)} annual)")
        logger.info(f"Period review needed: {session['periods_need_review']}")
        
        return {
            "session_id": session_id,
            "selected_sheets": sheet_selection,
            "period_analysis": period_analysis,
            "periods_need_review": session["periods_need_review"],
            "status": "periods_detected"
        }
        
    except Exception as e:
        logger.error(f"Sheet selection failed: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process sheet selection: {str(e)}"
        )

@router.get("/session/{session_id}/periods")
async def get_available_periods(session_id: str):
    """
    Get available periods from selected financial statement sheets
    """
    if session_id not in active_sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = active_sessions[session_id]
    
    periods = session.get("available_periods", [])
    selected_sheets = session.get("selected_sheets", {})
    
    return {
        "session_id": session_id,
        "available_periods": periods,
        "selected_sheets": selected_sheets,
        "default_period": periods[0] if periods else None
    }

@router.post("/session/{session_id}/approve-periods")
async def approve_detected_periods(session_id: str, approval: dict):
    """
    User approves the detected periods or requests template input
    
    Expected request body:
    {
        "approved": true/false,
        "custom_templates": [{"pattern": "...", "description": "..."}]  # if approved=false
    }
    """
    if session_id not in active_sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = active_sessions[session_id]
    
    try:
        if approval.get("approved", False):
            # User approved the detected periods
            period_analysis = session.get("period_analysis", {})
            periods_list = (
                period_analysis.get("annual_periods", []) + 
                period_analysis.get("quarterly_periods", []) + 
                period_analysis.get("other_periods", [])
            )
            
            session["available_periods"] = sorted(periods_list)
            session["periods_approved"] = True
            
            logger.info(f"User approved {len(periods_list)} detected periods")
            
            return {
                "session_id": session_id,
                "status": "periods_approved",
                "available_periods": session["available_periods"],
                "total_periods": len(periods_list)
            }
        else:
            # User wants to use custom templates
            custom_templates = approval.get("custom_templates", [])
            
            if not custom_templates:
                raise HTTPException(
                    status_code=400,
                    detail="Custom templates required when not approving detected periods"
                )
            
            # Use template parser to find periods with custom templates
            parser = UniversalExcelParser()
            old_file_path = session.get("old_file_path")
            selected_sheets = session.get("selected_sheets", {})
            
            if not old_file_path or not selected_sheets:
                raise HTTPException(status_code=400, detail="Session data incomplete")
            
            # Convert custom templates to PeriodTemplate objects
            from app.services.template_parser import PeriodTemplate
            template_objects = []
            
            for template_data in custom_templates:
                template_objects.append(PeriodTemplate(
                    name=template_data.get("name", "Custom Template"),
                    pattern=template_data["pattern"],
                    example=template_data.get("example", ""),
                    type=template_data.get("type", "quarterly")
                ))
            
            # Re-parse with custom templates
            old_statements = parser.parse_financial_statements(old_file_path, selected_sheets)
            
            # Use template-based detection for each sheet
            all_periods = set()
            enhanced_details = {}
            
            for statement_type, statement in old_statements.items():
                # Get the sheet for template parsing
                import openpyxl
                wb = openpyxl.load_workbook(old_file_path, data_only=False)
                sheet = wb[selected_sheets[statement_type]]
                
                # Find additional periods using templates
                template_results = parser.template_parser.find_periods_using_templates(
                    sheet, template_objects
                )
                
                # Combine standard and template results
                enhanced_periods = list(statement.periods)
                for template_name, template_matches in template_results.items():
                    for period, column in template_matches:
                        if period not in enhanced_periods:
                            enhanced_periods.append(period)
                
                all_periods.update(enhanced_periods)
                enhanced_details[statement_type] = {
                    "periods_found": len(enhanced_periods),
                    "original_periods": len(statement.periods),
                    "template_periods": len(enhanced_periods) - len(statement.periods)
                }
                
                wb.close()
            
            periods_list = sorted(list(all_periods))
            session["available_periods"] = periods_list
            session["periods_approved"] = True
            session["used_custom_templates"] = True
            
            logger.info(f"Custom templates found {len(periods_list)} total periods")
            
            return {
                "session_id": session_id,
                "status": "custom_templates_applied",
                "available_periods": periods_list,
                "total_periods": len(periods_list),
                "enhancement_details": enhanced_details,
                "templates_used": len(custom_templates)
            }
            
    except Exception as e:
        logger.error(f"Period approval failed: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process period approval: {str(e)}"
        )

@router.get("/session/{session_id}/template-hints")
async def get_template_hints(session_id: str):
    """
    Get hints and examples for creating custom period templates
    """
    return {
        "session_id": session_id,
        "template_hints": {
            "placeholders": {
                "{Q}": "Quarter number (1, 2, 3, 4)",
                "{YY}": "2-digit year (25 for 2025)",
                "{YYYY}": "4-digit year (2025)",
                "{WW}": "Week number (01-53)",
                "[E]": "Optional estimate suffix"
            },
            "examples": [
                {
                    "pattern": "FY{Q}Q{YY}[E]",
                    "description": "Fiscal year quarterly with optional estimate",
                    "generates": ["FY1Q25", "FY1Q25E", "FY2Q25", "FY2Q25E"]
                },
                {
                    "pattern": "{Q}Q{YY}[E]", 
                    "description": "Standard quarterly with optional estimate",
                    "generates": ["1Q25", "1Q25E", "2Q25", "2Q25E"]
                },
                {
                    "pattern": "Q{Q} {YYYY}",
                    "description": "Quarter with full year",
                    "generates": ["Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025"]
                },
                {
                    "pattern": "FY{YYYY}[E]",
                    "description": "Annual fiscal year with optional estimate", 
                    "generates": ["FY2025", "FY2025E", "FY2026", "FY2026E"]
                }
            ],
            "tips": [
                "Use square brackets [] for optional suffixes like estimates",
                "Templates will generate periods for years 1990-2030 by default",
                "Multiple templates can be combined to capture different formats",
                "Test your template pattern by looking at the 'generates' examples"
            ]
        }
    }

def _is_valid_file(filename: str) -> bool:
    """Check if the uploaded file has a valid extension"""
    if not filename:
        return False
    return Path(filename).suffix.lower() in settings.ALLOWED_EXTENSIONS

async def _save_upload_file(upload_file: UploadFile, destination: Path):
    """Save uploaded file to destination path"""
    try:
        with open(destination, "wb") as buffer:
            shutil.copyfileobj(upload_file.file, buffer)
    finally:
        upload_file.file.close()