"""
Universal Excel Parser Service
Handles any Excel financial model structure for variance analysis
"""
import openpyxl
import re
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
import logging
import dataclasses
from dataclasses import dataclass
from fuzzywuzzy import fuzz
from .formula_analyzer import FormulaAnalyzer, FormulaComponent, DrillDownResult
from .template_parser import TemplateBasedPeriodParser, PeriodTemplate


# Import additional classes from dual_parser for enhanced functionality
try:
    from app.models.financial import SheetType, FinancialModel
    from app.utils.excel_utils import detect_financial_keywords, find_period_header_row
except ImportError as e:
    print(f"Could not import some classes: {e}. Using fallbacks.")

logger = logging.getLogger(__name__)

@dataclass
class LineItem:
    name: str
    row_number: int
    values: Dict[str, float]  # period -> value mapping
    formula: Optional[str] = None
    dependencies: List[str] = dataclasses.field(default_factory=list)  # Cell references for drill-down

@dataclass
class FinancialStatement:
    sheet_name: str
    periods: List[str]  # Detected from column headers
    line_items: Dict[int, LineItem]  # row_number -> LineItem
    period_columns: Dict[str, int]  # period -> column_number mapping

@dataclass
class ModelComparison:
    old_model: Dict[str, FinancialStatement]  # statement_type -> statement
    new_model: Dict[str, FinancialStatement]
    selected_sheets: Dict[str, str]  # statement_type -> sheet_name
    selected_period: str
    variances: Dict[str, Any] = dataclasses.field(default_factory=dict)

class UniversalExcelParser:
    """Parse any Excel financial model for variance analysis"""
    
    def __init__(self):
        # Extensible regex patterns for period detection
        self.period_patterns = [
            r'Q(\d)\s?(\d{4})',           # Q1 2024, Q1 2024
            r'(\d)Q(\d{2,4})[E]?',        # 1Q24, 1Q2024, 1Q24E, 1Q2024E
            r'FY(\d)Q(\d{2,4})[E]?',      # FY1Q25, FY2Q24E (fiscal year quarterly)
            r'FY\s?(\d{4})[E]?',          # FY 2024, FY2024, FY2024E
            r'(\d{4})[E]?',               # 2024, 2024E
            r'(\w{3})\s?(\d{4})',         # Mar 2024, Mar2024
            r'(\d{1,2})/(\d{4})',         # 3/2024, 12/2024
            r'(\w{3})-(\d{2,4})',         # Mar-24, Mar-2024
            r'CY\s?(\d{4})[E]?',          # CY2024, CY 2024, CY2024E
            r'(\d{4})\s?(Actual|Estimate|Forecast|Budget)', # 2024 Actual, 2024E
            r'FY(\d)Q(\d{2})?[E]?',       # FY1Q, FY2Q, FY3Q, FY4Q (without year)
            r'(\d{4})-(\d{2})',           # 1998-53 format
        ]
        
        # Compile patterns for performance
        self.compiled_patterns = [re.compile(pattern, re.IGNORECASE) for pattern in self.period_patterns]
        
        # Initialize formula analyzer
        self.formula_analyzer = FormulaAnalyzer()
        
        # Initialize template-based parser
        self.template_parser = TemplateBasedPeriodParser()
    
    def parse_financial_statements(self, file_path: Path, selected_sheets: Dict[str, str]) -> Dict[str, FinancialStatement]:
        """
        Parse selected financial statement sheets from an Excel file
        
        Args:
            file_path: Path to Excel file
            selected_sheets: {'income_statement': 'IS Sheet Name', 'balance_sheet': 'BS Sheet Name', ...}
        
        Returns:
            Dictionary mapping statement types to parsed FinancialStatement objects
        """
        if not file_path.exists():
            error_msg = f"Excel file not found: {file_path}"
            logger.error(error_msg)
            raise FileNotFoundError(error_msg)
            
        if not selected_sheets:
            error_msg = "No sheets selected for parsing"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        logger.info(f"Parsing financial statements from {file_path}")
        logger.info(f"Selected sheets: {selected_sheets}")
        
        wb = None
        try:
            # Try to load workbook with comprehensive error handling
            try:
                wb = openpyxl.load_workbook(file_path, read_only=False, data_only=False)
                logger.info(f"Successfully loaded workbook with {len(wb.sheetnames)} sheets")
            except PermissionError as perm_error:
                error_msg = f"Permission denied accessing {file_path}. File may be open in Excel or access restricted."
                logger.error(error_msg)
                raise PermissionError(error_msg) from perm_error
            except Exception as load_error:
                error_msg = f"Failed to load Excel workbook {file_path}: {str(load_error)}"
                logger.error(error_msg)
                raise ValueError(error_msg) from load_error
            
            statements = {}
            sheets_found = []
            sheets_missing = []
            parsing_errors = []
            
            # Validate all requested sheets exist before parsing any
            for statement_type, sheet_name in selected_sheets.items():
                if sheet_name in wb.sheetnames:
                    sheets_found.append(f"{statement_type}:{sheet_name}")
                else:
                    sheets_missing.append(f"{statement_type}:{sheet_name}")
                    logger.error(f"Sheet '{sheet_name}' not found for {statement_type}. Available sheets: {wb.sheetnames}")
            
            if sheets_missing:
                error_msg = f"Missing sheets: {sheets_missing}. Available sheets: {wb.sheetnames}"
                logger.error(error_msg)
                raise ValueError(error_msg)
            
            logger.info(f"All requested sheets found: {sheets_found}")
            
            # Parse each sheet with individual error handling
            for statement_type, sheet_name in selected_sheets.items():
                try:
                    logger.info(f"Parsing {statement_type} from sheet: {sheet_name}")
                    statement = self._parse_sheet(wb[sheet_name], sheet_name, file_path)
                    
                    # Validate parsing results
                    if not statement.periods:
                        warning_msg = f"No periods found in {sheet_name} for {statement_type}"
                        logger.warning(warning_msg)
                        # Don't fail completely, but warn user
                    
                    if not statement.line_items:
                        warning_msg = f"No line items found in {sheet_name} for {statement_type}"
                        logger.warning(warning_msg)
                    
                    statements[statement_type] = statement
                    logger.info(f"Successfully parsed {statement_type}: {len(statement.periods)} periods, {len(statement.line_items)} line items")
                    
                except Exception as sheet_error:
                    error_msg = f"Failed to parse sheet {sheet_name} for {statement_type}: {str(sheet_error)}"
                    logger.error(error_msg)
                    parsing_errors.append(error_msg)
                    # Continue parsing other sheets rather than failing completely
            
            # Check if we successfully parsed any sheets
            if not statements:
                error_msg = f"Failed to parse any sheets. Errors: {parsing_errors}"
                logger.error(error_msg)
                raise RuntimeError(error_msg)
            
            if parsing_errors:
                logger.warning(f"Some sheets failed to parse: {parsing_errors}")
                # Continue with successfully parsed sheets
            
            logger.info(f"Parsing completed successfully: {len(statements)} statements parsed")
            return statements
            
        except Exception as e:
            error_msg = f"Failed to parse {file_path}: {str(e)}"
            logger.error(error_msg)
            # Re-raise with original exception context
            raise type(e)(error_msg) from e
            
        finally:
            # Ensure workbook is always closed
            if wb:
                try:
                    wb.close()
                    logger.debug("Workbook closed successfully")
                except Exception as close_error:
                    logger.warning(f"Error closing workbook: {close_error}")
    
    def _parse_sheet(self, sheet, sheet_name: str, file_path=None, user_templates: List[PeriodTemplate] = None) -> FinancialStatement:
        """Parse a single financial statement sheet"""
        
        # Step 1: Detect periods using enhanced template-based approach
        periods, period_columns = self._detect_periods_with_templates(sheet, file_path, user_templates)
        logger.info(f"Detected periods in {sheet_name}: {len(periods)} periods found")
        
        # Step 2: Extract line items and their values
        line_items = self._extract_line_items(sheet, periods, period_columns)
        logger.info(f"Extracted {len(line_items)} line items from {sheet_name}")
        
        return FinancialStatement(
            sheet_name=sheet_name,
            periods=periods,
            line_items=line_items,
            period_columns=period_columns
        )
    
    def _detect_periods(self, sheet, data_only_sheet=None) -> Tuple[List[str], Dict[str, int]]:
        """Detect periods from column headers using regex patterns"""
        
        periods = []
        period_columns = {}
        
        # Get sheet dimensions for logging and performance monitoring
        max_rows = sheet.max_row
        max_cols = sheet.max_column
        
        if max_rows is None or max_cols is None:
            error_msg = f"Unable to determine sheet dimensions for {sheet.title}"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        # Performance warning for very wide sheets
        if max_cols > 500:
            logger.warning(f"Sheet {sheet.title} has {max_cols} columns - period detection may be slow")
        
        # Calculate scanning parameters
        header_rows_to_scan = min(11, max_rows + 1)  # Rows 1-10
        columns_to_scan = max_cols + 1  # Scan ALL columns (no more 50-column limit)
        total_cells_to_scan = (header_rows_to_scan - 1) * (columns_to_scan - 1)
        
        logger.info(f"Scanning sheet {sheet.title}: {header_rows_to_scan-1} rows × {columns_to_scan-1} columns = {total_cells_to_scan} cells")
        
        try:
            cells_processed = 0
            cells_with_data = 0
            
            # Check first few rows for headers (usually rows 1-10, to catch various header structures)
            for row_num in range(1, header_rows_to_scan):
                for col_num in range(1, columns_to_scan):
                    try:
                        cell = sheet.cell(row=row_num, column=col_num)
                        cell_value = cell.value
                        cells_processed += 1
                        
                        if cell_value and isinstance(cell_value, str):
                            cells_with_data += 1
                            
                            # Handle Excel formulas that might generate period names
                            actual_value = self._get_cell_display_value(sheet, cell, cell_value, data_only_sheet)
                            
                            # Try to match against all period patterns
                            period = self._match_period_pattern(actual_value.strip())
                            if period and period not in periods:
                                periods.append(period)
                                period_columns[period] = col_num
                                cell_coordinate = cell.coordinate
                                logger.debug(f"Found period '{period}' in cell {cell_coordinate}")
                                
                    except Exception as cell_error:
                        # Don't let individual cell errors stop the entire process
                        cell_coordinate = f"{self._number_to_column_letter(col_num)}{row_num}"
                        logger.warning(f"Error reading cell {cell_coordinate}: {cell_error}")
                        continue
            
            # Log scanning results
            logger.info(f"Period detection completed: {cells_processed} cells processed, {cells_with_data} contained data, {len(periods)} periods found")
            
            # Validate we found reasonable results
            if cells_processed == 0:
                error_msg = f"No cells were processed in sheet {sheet.title} - this indicates a serious parsing error"
                logger.error(error_msg)
                raise RuntimeError(error_msg)
                
            if cells_with_data == 0:
                logger.warning(f"No cells with data found in header rows of sheet {sheet.title}")
                
        except Exception as scanning_error:
            error_msg = f"Failed to scan sheet {sheet.title} for periods: {str(scanning_error)}"
            logger.error(error_msg)
            raise RuntimeError(error_msg) from scanning_error
        
        # Return periods in Excel column order (no sorting)
        logger.debug(f"Returning {len(periods)} periods in Excel column order")
        return periods, period_columns
    
    def _detect_periods_alternative(self, file_path, sheet_name: str) -> Tuple[List[str], Dict[str, int]]:
        """
        Alternative period detection using direct data_only approach
        This bypasses the formula evaluation issues in the main parser
        """
        periods = []
        period_columns = {}
        
        try:
            # Method 1: Try data_only=True (this is what works!)
            logger.info(f"Alternative detection: Attempting to load {file_path} (type: {type(file_path)})")
            wb_data = openpyxl.load_workbook(str(file_path), data_only=True)
            sheet_data = wb_data[sheet_name]
            
            logger.info(f"Alternative detection: Successfully loaded {sheet_name} with data_only=True")
            
            # Scan first 10 rows across all columns
            for row_num in range(1, 11):
                for col_num in range(1, sheet_data.max_column + 1):
                    try:
                        cell = sheet_data.cell(row=row_num, column=col_num)
                        value = cell.value
                        
                        if value and isinstance(value, str):
                            period = self._match_period_pattern(value.strip())
                            if period and period not in periods:
                                periods.append(period)
                                period_columns[period] = col_num
                                logger.debug(f"Alternative detection found: {period} at column {col_num}")
                    
                    except Exception:
                        continue
            
            wb_data.close()
            logger.info(f"Alternative detection found {len(periods)} periods")
            
            return periods, period_columns
            
        except Exception as e:
            logger.warning(f"Alternative detection failed: {e}")
            return [], {}

    def _detect_periods_with_templates(self, sheet, file_path=None, user_templates: List[PeriodTemplate] = None) -> Tuple[List[str], Dict[str, int]]:
        """
        Enhanced period detection using template-based approach
        
        Args:
            sheet: Excel worksheet
            user_templates: Optional list of user-provided templates
            
        Returns:
            Tuple of (periods_list, period_to_column_mapping)
        """
        
        # Step 1: Try alternative detection approach first (bypasses formula evaluation issues)
        if file_path is None:
            try:
                file_path = sheet.parent.path
            except:
                file_path = None
        sheet_name = sheet.title
        
        alt_periods, alt_columns = self._detect_periods_alternative(file_path, sheet_name)
        
        if len(alt_periods) >= 50:  # If alternative approach finds good results, use it
            logger.info(f"Alternative detection successful: {len(alt_periods)} periods found")
            return alt_periods, alt_columns
        
        # Step 2: Fall back to standard detection if alternative doesn't find enough
        logger.info(f"Alternative detection found only {len(alt_periods)} periods, trying standard approach")
        
        # Load data_only workbook for formula evaluation
        data_only_wb = None
        data_only_sheet = None
        try:
            # Try to load workbook in data_only mode for formula evaluation
            data_only_wb = openpyxl.load_workbook(str(file_path), data_only=True)
            data_only_sheet = data_only_wb[sheet_name]
            logger.debug("Loaded data_only workbook for formula evaluation")
        except Exception as e:
            logger.warning(f"Could not load data_only workbook: {e}")
        
        # Try standard detection with formula evaluation
        standard_periods, standard_columns = self._detect_periods(sheet, data_only_sheet)
        
        # Clean up data_only workbook
        if data_only_wb:
            try:
                data_only_wb.close()
            except:
                pass
        
        # Merge alternative and standard results
        merged_periods = list(alt_periods)
        merged_columns = dict(alt_columns)
        
        for period in standard_periods:
            if period not in merged_periods:
                merged_periods.append(period)
                merged_columns[period] = standard_columns.get(period, 0)
        
        logger.info(f"Combined detection found {len(merged_periods)} periods")
        
        # Step 3: If we found fewer than expected, use template approach
        if len(merged_periods) < 70:  # Threshold for "probably missing periods"
            
            logger.info("Low period count detected, applying template-based enhancement")
            
            # Suggest templates from found periods if none provided
            templates = user_templates
            if not templates:
                templates = self.template_parser.suggest_templates_from_samples(merged_periods)
                logger.info(f"Auto-suggested {len(templates)} templates")
            
            # Use templates to find additional periods
            if templates:
                # Use the same header rows as the main parser
                header_rows_to_scan = min(11, sheet.max_row + 1)
                header_rows = list(range(1, header_rows_to_scan))
                found_by_template = self.template_parser.find_periods_using_templates(sheet, templates, header_rows)
                
                # Merge template results with merged results
                enhanced_periods = list(merged_periods)
                enhanced_columns = dict(merged_columns)
                
                for template_name, template_matches in found_by_template.items():
                    for period, column in template_matches:
                        if period not in enhanced_periods:
                            enhanced_periods.append(period)
                            enhanced_columns[period] = column
                
                logger.info(f"Template enhancement found {len(enhanced_periods) - len(merged_periods)} additional periods")
                
                # Sort by column number to maintain Excel order
                enhanced_periods.sort(key=lambda p: enhanced_columns.get(p, 0))
                
                return enhanced_periods, enhanced_columns
        
        # Return merged results if template enhancement not needed/available
        return merged_periods, merged_columns
    
    def _match_period_pattern(self, text: str) -> Optional[str]:
        """Match text against period patterns and return normalized period"""
        
        # Skip obvious phone numbers or long numeric strings
        if re.match(r'^\d{3}-\d{3}-\d{4}$', text) or re.match(r'^\d{10,}$', text):
            return None
            
        # Skip if it contains obvious non-period content
        if any(word in text.lower() for word in ['phone', 'tel', 'fax', 'contact']):
            return None
        
        # Skip Excel formulas that look like period generators (these need to be evaluated)
        if text.startswith('=') and any(keyword in text for keyword in ['FY', 'Q', 'RIGHT', 'LEFT']):
            logger.debug(f"Skipping formula for later evaluation: {text}")
            return None
        
        for pattern in self.compiled_patterns:
            match = pattern.search(text)
            if match:
                matched_text = match.group().strip()
                # Additional validation: don't match 4-digit years that are too high/low
                if re.match(r'^\d{4}$', matched_text):
                    year = int(matched_text)
                    if year < 1990 or year > 2050:  # Reasonable business date range
                        continue
                # Don't filter out the full matched text - return the original text, not just the matched group
                return text.strip()
        
        return None
    
    def _normalize_period(self, period_text: str) -> str:
        """Normalize period text to consistent format"""
        # Keep original for now, but could standardize formats here
        # e.g., convert "1Q24" to "Q1 2024"
        return period_text
    
    def _sort_periods_chronologically(self, periods: List[str]) -> List[str]:
        """Sort periods in chronological order"""
        
        def period_sort_key(period: str) -> Tuple[int, int]:
            """Extract year and quarter/month for sorting"""
            try:
                # Extract year - handle both 4-digit and 2-digit years
                year = 9999  # Default for unrecognized
                quarter = 0  # Default for annual
                
                # Handle formats like "3Q25E", "1Q24", etc.
                quarter_year_match = re.search(r'(\d)Q(\d{2})E?', period)
                if quarter_year_match:
                    quarter = int(quarter_year_match.group(1))
                    year = 2000 + int(quarter_year_match.group(2))
                    return (year, quarter)
                
                # Handle formats like "Q1 2024", "Q3 25"
                year_quarter_match = re.search(r'Q(\d)\s?(\d{2,4})', period)
                if year_quarter_match:
                    quarter = int(year_quarter_match.group(1))
                    year_str = year_quarter_match.group(2)
                    year = int(year_str) if len(year_str) == 4 else 2000 + int(year_str)
                    return (year, quarter)
                
                # Handle 4-digit years
                year_match_4 = re.search(r'(\d{4})', period)
                if year_match_4:
                    year = int(year_match_4.group(1))
                
                # Check for quarterly indicators without year context
                quarter_only_match = re.search(r'Q(\d)', period)
                if quarter_only_match:
                    quarter = int(quarter_only_match.group(1))
                    return (year, quarter)
                
                # Extract month
                month_match = re.search(r'(\d{1,2})/', period)
                if month_match:
                    return (year, int(month_match.group(1)))
                
                # Annual periods (no quarter)
                return (year, 0)
                
            except (ValueError, AttributeError):
                return (9999, 99)  # Put unrecognized formats at the end
        
        return sorted(periods, key=period_sort_key)
    
    def _extract_line_items(self, sheet, periods: List[str], period_columns: Dict[str, int]) -> Dict[int, LineItem]:
        """Extract line items and their values from the sheet"""
        line_items = {}
        try:
            period_header_row = find_period_header_row(sheet)
            print(f"DEBUG: Detected period header row at Excel row {period_header_row} in sheet '{sheet.title}'")
            start_row = period_header_row + 1
            logger.info(f"Detected period header row at {period_header_row}, starting line item extraction from row {start_row}")
        except Exception as e:
            logger.warning(f"Could not detect period header row: {e}, defaulting to row 6")
            start_row = 6

        for row_num in range(start_row, sheet.max_row + 1):
            # Check first few columns for line item names
            line_item_name = None
            for col_num in range(1, min(6, sheet.max_column + 1)):
                cell_value = sheet.cell(row=row_num, column=col_num).value
                if cell_value and isinstance(cell_value, str) and cell_value.strip():
                    # Skip if it looks like a period header
                    if not self._match_period_pattern(cell_value):
                        line_item_name = cell_value.strip()
                        break

            if not line_item_name:
                continue

            # Extract values for each period
            values = {}
            formulas = {}

            for period in periods:
                col_num = period_columns.get(period)
                if col_num:
                    cell = sheet.cell(row=row_num, column=col_num)

                    # Get value (with data_only=False, we get formulas)
                    if cell.data_type == 'f':  # Formula
                        formulas[period] = cell.value
                        # Try to get calculated value
                        try:
                            # Create a data_only workbook to get calculated values
                            wb_data = openpyxl.load_workbook(sheet.parent.path, data_only=True)
                            calc_cell = wb_data[sheet.title].cell(row=row_num, column=col_num)
                            if isinstance(calc_cell.value, (int, float)):
                                values[period] = float(calc_cell.value)
                            wb_data.close()
                        except:
                            # If we can't get calculated value, set to 0
                            values[period] = 0.0
                    elif isinstance(cell.value, (int, float)):
                        values[period] = float(cell.value)

            # Only add line items that have at least one non-zero value OR formulas
            has_meaningful_data = (values and any(abs(v) > 0.001 for v in values.values())) or bool(formulas)
            if has_meaningful_data:
                # Use row number as key - this naturally handles duplicates and preserves Excel structure
                # Get the main formula (from the first period with a formula)
                main_formula = next((f for f in formulas.values() if f), None)

                line_items[row_num] = LineItem(
                    name=line_item_name,
                    row_number=row_num,
                    values=values,
                    formula=main_formula,
                    dependencies=self._extract_formula_dependencies(main_formula) if main_formula else []
                )

        return line_items
    
    def _extract_formula_dependencies(self, formula: str) -> List[str]:
        """Extract cell references from Excel formula"""
        if not formula:
            return []
        
        # Simple regex to find cell references like A1, $B$2, Sheet1!C3
        cell_pattern = r'(?:\w+!)?(?:\$)?[A-Z]+(?:\$)?[0-9]+(?::\$?[A-Z]+\$?[0-9]+)?'
        dependencies = re.findall(cell_pattern, formula)
        
        return list(set(dependencies))  # Remove duplicates
    
    def match_line_items(self, old_statement: FinancialStatement, new_statement: FinancialStatement) -> List[Tuple[str, str]]:
        """
        Match line items between old and new statements using fuzzy matching
        
        Returns:
            List of tuples (old_line_item_name, new_line_item_name)
        """
        matches = []
        old_items = list(old_statement.line_items.keys())
        new_items = list(new_statement.line_items.keys())
        
        # First pass: exact matches
        exact_matches = []
        for old_item in old_items[:]:
            if old_item in new_items:
                matches.append((old_item, old_item))
                exact_matches.append(old_item)
                new_items.remove(old_item)
        
        # Remove exact matches from old_items
        for item in exact_matches:
            old_items.remove(item)
        
        # Second pass: fuzzy matching for remaining items
        similarity_threshold = 85  # Adjust as needed
        
        for old_item in old_items:
            best_match = None
            best_score = 0
            
            for new_item in new_items:
                score = fuzz.ratio(old_item.lower(), new_item.lower())
                if score > best_score and score >= similarity_threshold:
                    best_score = score
                    best_match = new_item
            
            if best_match:
                matches.append((old_item, best_match))
                new_items.remove(best_match)
        
        logger.info(f"Matched {len(matches)} line items between statements")
        return matches
    
    def calculate_variances(self, old_statement: FinancialStatement, new_statement: FinancialStatement, 
                          period: str, matches: List[Tuple[str, str]]) -> Dict[str, Any]:
        """Calculate variances for matched line items"""
        
        variances = {}
        
        for old_name, new_name in matches:
            old_item = old_statement.line_items.get(old_name)
            new_item = new_statement.line_items.get(new_name)
            
            if old_item and new_item:
                old_value = old_item.values.get(period, 0.0)
                new_value = new_item.values.get(period, 0.0)
                
                absolute_variance = new_value - old_value
                percentage_variance = (absolute_variance / old_value * 100) if old_value != 0 else 0
                
                variances[old_name] = {
                    'line_item_name': old_name,
                    'matched_with': new_name if old_name != new_name else None,
                    'old_value': old_value,
                    'new_value': new_value,
                    'absolute_variance': absolute_variance,
                    'percentage_variance': percentage_variance,
                    'has_formula': bool(old_item.formula or new_item.formula),
                    'drill_down_available': len(old_item.dependencies) > 0 or len(new_item.dependencies) > 0
                }
        
        return variances
    
    def drill_down_variance(self, old_file_path: Path, new_file_path: Path,
                          sheet_name: str, line_item_name: str, period: str) -> Optional[DrillDownResult]:
        """
        Drill down into a specific line item to show component variances
        
        Args:
            old_file_path: Path to old Excel file
            new_file_path: Path to new Excel file
            sheet_name: Name of the sheet containing the line item
            line_item_name: Name of the line item to drill down
            period: Period to analyze
            
        Returns:
            DrillDownResult with component breakdown
        """
        try:
            logger.info(f"Drilling down into {line_item_name} on {sheet_name} for period {period}")
            
            # Parse both files to find the line item
            selected_sheets = {"target": sheet_name}
            
            old_statements = self.parse_financial_statements(old_file_path, selected_sheets)
            new_statements = self.parse_financial_statements(new_file_path, selected_sheets)
            
            if "target" not in old_statements or "target" not in new_statements:
                logger.error("Could not parse target sheet")
                return None
            
            old_stmt = old_statements["target"]
            new_stmt = new_statements["target"]
            
            # Find the line item in both statements
            old_item = old_stmt.line_items.get(line_item_name)
            new_item = new_stmt.line_items.get(line_item_name)
            
            if not old_item or not new_item:
                logger.error(f"Line item '{line_item_name}' not found in both models")
                return None
            
            # Check if the item has a formula
            if not old_item.formula and not new_item.formula:
                logger.info(f"Line item '{line_item_name}' has no formula - cannot drill down")
                return None
            
            # Use the formula from whichever model has it
            source_formula = old_item.formula or new_item.formula
            source_row = old_item.row_number
            
            # Build dependency trees for both old and new
            old_cell_address = f"{self._get_column_from_period(old_stmt, period)}{source_row}"
            new_cell_address = f"{self._get_column_from_period(new_stmt, period)}{source_row}"
            
            logger.info(f"Analyzing formula at {sheet_name}!{old_cell_address}")
            
            old_component = self.formula_analyzer.build_dependency_tree(
                old_file_path, sheet_name, old_cell_address, max_depth=2
            )
            
            new_component = self.formula_analyzer.build_dependency_tree(
                new_file_path, sheet_name, new_cell_address, max_depth=2
            )
            
            if not old_component or not new_component:
                logger.error("Could not build dependency trees")
                return None
            
            # Calculate variance attribution
            drill_down_result = self.formula_analyzer.calculate_variance_attribution(
                old_component, new_component
            )
            
            logger.info(f"Drill-down completed: {len(drill_down_result.components)} components found")
            return drill_down_result
            
        except Exception as e:
            logger.error(f"Drill-down failed: {str(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            return None
    
    def _get_column_from_period(self, statement: FinancialStatement, period: str) -> str:
        """Get the Excel column letter for a specific period"""
        if period in statement.period_columns:
            col_num = statement.period_columns[period]
            return self._number_to_column_letter(col_num)
        return "A"  # Fallback
    
    def _number_to_column_letter(self, column_number: int) -> str:
        """Convert number to Excel column letter"""
        result = ""
        while column_number > 0:
            column_number -= 1
            result = chr(column_number % 26 + ord('A')) + result
            column_number //= 26
        return result
    
    def _get_cell_display_value(self, sheet, cell, cell_value: str, data_only_sheet=None) -> str:
        """
        Get the display value of a cell, evaluating formulas if possible
        
        Args:
            sheet: The Excel worksheet
            cell: The openpyxl cell object
            cell_value: The raw cell value
            
        Returns:
            The display value (formula result if applicable)
        """
        # If it's not a formula, return as-is
        if not cell_value.startswith('='):
            return cell_value
        
        # For formulas, try to get the calculated value using data_only sheet
        if data_only_sheet:
            try:
                data_only_cell = data_only_sheet.cell(row=cell.row, column=cell.column)
                if data_only_cell.value is not None:
                    calculated_value = str(data_only_cell.value)
                    logger.debug(f"Formula {cell_value} evaluated to: {calculated_value}")
                    return calculated_value
                    
            except Exception as eval_error:
                logger.debug(f"Could not evaluate formula {cell_value}: {eval_error}")
        
        # If formula evaluation fails, try simple text substitution for common patterns
        if 'FY' in cell_value and 'Q' in cell_value and 'RIGHT' in cell_value:
            # Pattern like ="FY1Q"&RIGHT(BJ6,2) - try to extract the quarter
            quarter_match = re.search(r'"FY(\d)Q"', cell_value)
            if quarter_match:
                quarter = quarter_match.group(1)
                # Try to guess year from nearby cells or use a recent year
                estimated_period = f"FY{quarter}Q25"  # Default to 2025
                logger.debug(f"Estimated formula result: {cell_value} -> {estimated_period}")
                return estimated_period
        
        # If all else fails, return the original formula
        return cell_value
    
    def get_drill_down_preview(self, file_path: Path, sheet_name: str, 
                              line_item_name: str) -> Dict[str, Any]:
        """
        Get a preview of what drilling down would show (for UI hints)
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet containing the line item
            line_item_name: Line item to preview
            
        Returns:
            Dictionary with drill-down preview information
        """
        try:
            selected_sheets = {"target": sheet_name}
            statements = self.parse_financial_statements(file_path, selected_sheets)
            
            if "target" not in statements:
                return {"can_drill_down": False, "reason": "Sheet not found"}
            
            stmt = statements["target"]
            item = stmt.line_items.get(line_item_name)
            
            if not item:
                return {"can_drill_down": False, "reason": "Line item not found"}
            
            if not item.formula:
                return {"can_drill_down": False, "reason": "No formula found"}
            
            # Analyze formula complexity
            complexity = self.formula_analyzer.analyze_formula_complexity(item.formula)
            
            return {
                "can_drill_down": complexity["can_drill_down"],
                "complexity": complexity["complexity"],
                "estimated_components": complexity["reference_count"],
                "has_cross_sheet_refs": complexity["has_cross_sheet_refs"],
                "has_external_refs": complexity["has_external_refs"],
                "main_function": complexity["main_function"],
                "estimated_depth": complexity["estimated_depth"]
            }
            
        except Exception as e:
            logger.warning(f"Could not preview drill-down for {line_item_name}: {e}")
            return {"can_drill_down": False, "reason": "Analysis error"}
    
    
    def _detect_sheet_type_enhanced(self, sheet_name: str, sheet_info: Dict[str, Any] = None) -> str:
        """
        Enhanced sheet type detection based on name and content analysis
        Copied from dual_parser for better compatibility
        """
        try:
            # First check sheet name
            sheet_name_lower = sheet_name.lower()
            
            # Income Statement keywords
            if any(keyword in sheet_name_lower for keyword in 
                   ['income', 'p&l', 'profit', 'loss', 'revenue', 'sales', 'earnings']):
                return "income_statement"
            
            # Balance Sheet keywords
            elif any(keyword in sheet_name_lower for keyword in 
                    ['balance', 'sheet', 'assets', 'liabilities', 'equity']):
                return "balance_sheet"
            
            # Cash Flow keywords
            elif any(keyword in sheet_name_lower for keyword in 
                    ['cash', 'flow', 'operating', 'investing', 'financing']):
                return "cash_flow"
            
            # If content analysis is available, use it
            if sheet_info:
                try:
                    sample_data = sheet_info.get('sample_data', [])
                    all_text = []
                    
                    for row_data in sample_data:
                        for cell_data in row_data:
                            value = cell_data.get('value')
                            if isinstance(value, str):
                                all_text.append(value)
                    
                    combined_text = ' '.join(all_text).lower()
                    
                    # Simple keyword scoring
                    income_keywords = ['revenue', 'sales', 'income', 'profit', 'loss', 'earnings', 'expense']
                    balance_keywords = ['assets', 'liabilities', 'equity', 'capital', 'retained', 'current']
                    cashflow_keywords = ['operating', 'investing', 'financing', 'cash', 'flow', 'payment']
                    
                    income_score = sum(1 for keyword in income_keywords if keyword in combined_text)
                    balance_score = sum(1 for keyword in balance_keywords if keyword in combined_text)
                    cashflow_score = sum(1 for keyword in cashflow_keywords if keyword in combined_text)
                    
                    if income_score > balance_score and income_score > cashflow_score:
                        return "income_statement"
                    elif balance_score > cashflow_score:
                        return "balance_sheet"
                    elif cashflow_score > 0:
                        return "cash_flow"
                        
                except Exception as e:
                    logger.debug(f"Content analysis failed for {sheet_name}: {e}")
            
            return "unknown"
            
        except Exception as e:
            logger.warning(f"Sheet type detection failed for {sheet_name}: {e}")
            return "unknown"
    
    def _consolidate_periods_enhanced(self, all_periods: List[str]) -> List[str]:
        """
        Enhanced period consolidation preserving Excel column order
        Copied from dual_parser for better compatibility
        """
        logger.info(f"📊 Consolidating {len(all_periods)} periods from all sheets")
        
        try:
            # Simple deduplication by name while preserving order
            seen_names = set()
            consolidated = []
            
            for period in all_periods:
                if isinstance(period, str) and period not in seen_names:
                    consolidated.append(period)
                    seen_names.add(period)
                    logger.debug(f"  📋 Added period: '{period}'")
            
            logger.info(f"✅ Consolidated to {len(consolidated)} unique periods in Excel order")
            return consolidated
            
        except Exception as e:
            logger.error(f"Period consolidation failed: {e}")
            return all_periods if isinstance(all_periods, list) else []