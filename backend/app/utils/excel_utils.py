"""
Excel utility functions for parsing and analysis
"""
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from typing import Dict, List, Tuple, Optional, Any
from pathlib import Path
import re
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

class ExcelReader:
    """Utility class for reading and analyzing Excel files"""
    
    def __init__(self, file_path: Path):
        self.file_path = file_path
        self.workbook = None
        self.sheets_info = {}
    
    def load_workbook(self) -> openpyxl.Workbook:
        """Load the Excel workbook"""
        try:
            # Use read_only mode for better performance with large files
            self.workbook = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
            logger.info(f"Loaded workbook: {self.file_path.name}")
            return self.workbook
        except Exception as e:
            logger.error(f"Failed to load workbook {self.file_path}: {str(e)}")
            raise
    
    def get_sheet_names(self) -> List[str]:
        """Get all sheet names in the workbook"""
        if not self.workbook:
            self.load_workbook()
        return self.workbook.sheetnames
    
    def analyze_sheet_content(self, sheet_name: str, max_rows: int = 50, max_cols: int = 20) -> Dict[str, Any]:
        """Analyze the content of a specific sheet"""
        if not self.workbook:
            self.load_workbook()
            
        sheet = self.workbook[sheet_name]
        
        # Get basic sheet info - limit dimensions for performance
        actual_max_row = min(sheet.max_row or 0, max_rows)
        actual_max_col = min(sheet.max_column or 0, max_cols)
        
        info = {
            'name': sheet_name,
            'max_row': sheet.max_row,
            'max_column': sheet.max_column,
            'analyzed_rows': actual_max_row,
            'analyzed_columns': actual_max_col,
            'has_data': sheet.max_row and sheet.max_row > 1,
        }
        
        # Analyze sample rows to identify structure
        sample_rows = []
        for row in range(1, actual_max_row + 1):
            row_data = []
            for col in range(1, actual_max_col + 1):
                cell = sheet.cell(row=row, column=col)
                # Get coordinate manually since read-only mode might not have it
                coord = f"{get_column_letter(col)}{row}"
                row_data.append({
                    'value': cell.value,
                    'formula': None,  # Formulas not available in data_only mode
                    'data_type': 'n' if isinstance(cell.value, (int, float)) else 's' if cell.value else None,
                    'coordinate': coord
                })
            sample_rows.append(row_data)
        
        info['sample_data'] = sample_rows
        return info
    
    def sheet_to_dataframe(self, sheet_name: str, header_row: int = 0) -> pd.DataFrame:
        """Convert sheet to pandas DataFrame"""
        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=header_row)
            return df
        except Exception as e:
            logger.warning(f"Failed to convert sheet {sheet_name} to DataFrame: {str(e)}")
            # Fallback: read without headers
            try:
                df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=None)
                return df
            except Exception as e2:
                logger.error(f"Complete failure reading sheet {sheet_name}: {str(e2)}")
                return pd.DataFrame()

def detect_financial_keywords(text: str) -> Dict[str, int]:
    """
    Detect financial keywords in text and return match scores
    """
    if not isinstance(text, str):
        return {}
    
    text_lower = text.lower()
    
    keyword_groups = {
        'income_statement': [
            'income', 'profit', 'loss', 'p&l', 'revenue', 'sales', 'ebitda', 'ebit',
            'operating', 'gross profit', 'net income', 'earnings', 'margin'
        ],
        'balance_sheet': [
            'balance', 'sheet', 'assets', 'liabilities', 'equity', 'cash', 'debt',
            'current assets', 'fixed assets', 'retained earnings', 'stockholder'
        ],
        'cash_flow': [
            'cash flow', 'operating cash', 'investing', 'financing', 'capex',
            'free cash flow', 'working capital', 'depreciation'
        ]
    }
    
    scores = {}
    for category, keywords in keyword_groups.items():
        score = sum(1 for keyword in keywords if keyword in text_lower)
        scores[category] = score
    
    return scores

def detect_period_patterns(text_list: List[str]) -> List[Dict[str, Any]]:
    """
    Detect date/period patterns in a list of text values
    """
    periods = []
    
    # Common period patterns
    patterns = {
        'quarterly': [
            r'Q[1-4]\s*20\d{2}',  # Q1 2024
            r'20\d{2}\s*Q[1-4]',  # 2024 Q1
            r'[1-4]Q\d{2}',       # 1Q24
        ],
        'yearly': [
            r'FY\s*20\d{2}',      # FY 2024
            r'20\d{2}',           # 2024
            r'CY\s*20\d{2}',      # CY 2024
        ],
        'monthly': [
            r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s*20\d{2}',  # Jan 2024
            r'20\d{2}-(0[1-9]|1[0-2])',  # 2024-01
        ]
    }
    
    for text in text_list:
        if not isinstance(text, str):
            continue
            
        for period_type, pattern_list in patterns.items():
            for pattern in pattern_list:
                matches = re.findall(pattern, text, re.IGNORECASE)
                for match in matches:
                    periods.append({
                        'text': match,
                        'type': period_type,
                        'original': text,
                        'pattern': pattern
                    })
    
    return periods

def extract_numeric_values(sheet_data: List[List[Dict]]) -> List[Tuple[str, float, str]]:
    """
    Extract numeric values with their coordinates from sheet data
    Returns list of (coordinate, value, data_type) tuples
    """
    numeric_values = []
    
    for row_idx, row_data in enumerate(sheet_data):
        for cell_data in row_data:
            value = cell_data.get('value')
            coordinate = cell_data.get('coordinate', '')
            data_type = cell_data.get('data_type', '')
            
            # Check if value is numeric
            if isinstance(value, (int, float)) and value != 0:
                numeric_values.append((coordinate, float(value), data_type))
            elif isinstance(value, str):
                # Try to parse string as number
                clean_value = re.sub(r'[^\d.-]', '', value)
                try:
                    if clean_value:
                        num_value = float(clean_value)
                        if num_value != 0:
                            numeric_values.append((coordinate, num_value, 'parsed'))
                except ValueError:
                    continue
    
    return numeric_values

def analyze_cell_relationships(sheet) -> Dict[str, List[str]]:
    """
    Analyze formula relationships in a sheet
    Returns mapping of cells to their dependencies
    """
    relationships = {}
    
    for row in sheet.iter_rows():
        for cell in row:
            if cell.data_type == 'f' and cell.value:  # Formula cell
                formula = cell.value
                # Extract cell references from formula
                cell_refs = re.findall(r'[A-Z]+\d+', formula)
                if cell_refs:
                    relationships[cell.coordinate] = cell_refs
    
    return relationships

def identify_hard_coded_values(sheet, relationships: Dict[str, List[str]]) -> List[str]:
    """
    Identify cells that contain hard-coded values (not formulas)
    """
    hard_coded = []
    
    for row in sheet.iter_rows():
        for cell in row:
            # Skip empty cells
            if cell.value is None:
                continue
                
            # If it's not a formula and has a numeric value, it might be hard-coded
            if cell.data_type != 'f' and isinstance(cell.value, (int, float)):
                # Additional check: not referenced by other cells
                is_referenced = any(cell.coordinate in deps for deps in relationships.values())
                if not is_referenced and cell.value != 0:
                    hard_coded.append(cell.coordinate)
    
    return hard_coded

def find_period_header_row(sheet) -> int:
    """
    Find the row that contains period headers in Excel file
    Simple approach: scan first 10 rows for period-like patterns
    """
    logger.info(f"🔍 Searching for period header row in sheet: {sheet.title}")
    
    for row_idx in range(1, min(11, sheet.max_row + 1)):  # Check first 10 rows
        row_data = [cell.value for cell in sheet[row_idx]]
        
        if looks_like_period_header(row_data):
            logger.info(f"✅ Found period header row at: {row_idx}")
            return row_idx
    
    logger.warning(f"⚠️ No period header row found in first 10 rows of {sheet.title}")
    raise ValueError(f"No period header row found in sheet: {sheet.title}")

def looks_like_period_header(row_data: List) -> bool:
    """
    Simple check: contains numbers, dates, or quarter patterns
    Must have at least 3 period-like values to qualify
    """
    period_indicators = 0
    
    for cell in row_data[1:]:  # Skip first column (usually line item names)
        if cell and isinstance(cell, (str, int, float)):
            cell_str = str(cell)
            # Look for period indicators: years, quarters, dates
            if any(pattern in cell_str for pattern in ['Q', '20', '19', '1Q', '2Q', '3Q', '4Q']):
                period_indicators += 1
            # Also check for 4-digit years
            elif len(cell_str) == 4 and cell_str.isdigit() and cell_str.startswith(('19', '20')):
                period_indicators += 1
    
    return period_indicators >= 3  # At least 3 period-like values

def extract_periods_from_row(sheet, header_row_idx: int, sheet_name: str) -> List['Period']:
    """
    Extract periods in exact column order from header row
    No normalization - use exact Excel text
    """
    from app.models.financial import Period, PeriodType
    
    logger.info(f"📊 Extracting periods from row {header_row_idx} in {sheet_name}")
    
    periods = []
    row = sheet[header_row_idx]
    
    for col_idx, cell in enumerate(row[1:], 1):  # Skip first column, start counting from 1
        if cell.value:
            period_name = str(cell.value).strip()
            if period_name:  # Only add non-empty periods
                period = Period(
                    name=period_name,
                    column_index=col_idx,
                    sheet_name=sheet_name,
                    row_index=header_row_idx,
                    period_type=PeriodType.FORECAST  # Default
                )
                periods.append(period)
                logger.debug(f"  📋 Period: '{period_name}' at column {col_idx}")
    
    logger.info(f"✅ Extracted {len(periods)} periods from {sheet_name}")
    return periods

def is_valid_line_item(first_cell, row, period_columns: List[int]) -> bool:
    """
    Check if row is a valid line item (DEPRECATED - use is_potential_line_item)
    Must have text in first column and at least one numeric value in period columns
    """
    if not first_cell or not isinstance(first_cell, str):
        return False
    
    if is_empty_or_formatting(first_cell):
        return False
    
    # Must have at least one numeric value in period columns
    for col_idx in period_columns:
        if col_idx < len(row):
            cell_value = row[col_idx].value if hasattr(row[col_idx], 'value') else row[col_idx]
            if isinstance(cell_value, (int, float)):
                return True
    
    return False


# REMOVED: is_potential_line_item function - was only used by removed extract_line_items_from_sheet

def is_pure_formatting(text: str) -> bool:
    """
    Check if text is just formatting characters
    More specific than is_empty_or_formatting
    """
    if not text or not text.strip():
        return True
    
    cleaned = text.strip()
    
    # Common formatting patterns
    formatting_chars = set(' -_=()[]{}|\\/:.,;')
    
    # If all characters are formatting, it's pure formatting
    if all(c in formatting_chars for c in cleaned):
        return True
    
    # Common formatting patterns
    formatting_patterns = [
        '---', '___', '===', '***', '...', 
        '----', '____', '====', '()', '[]', '{}'
    ]
    
    if any(pattern in cleaned for pattern in formatting_patterns):
        return True
    
    return False

def is_empty_row(row, period_columns: List[int]) -> bool:
    """
    Check if row is completely empty across first column and period columns
    """
    # Check first column
    first_cell = row[0].value
    if first_cell and str(first_cell).strip():
        return False
    
    # Check period columns
    for col_idx in period_columns:
        if col_idx < len(row):
            cell_value = row[col_idx].value
            if cell_value is not None:
                return False
    
    return True

# REMOVED: extract_any_numeric_value function - was only used by removed extract_line_items_from_sheet


def should_stop_processing(current_row_idx: int, empty_row_count: int, line_item_name: str, statement_type: str) -> bool:
    """
    Legacy function - replaced by new logic in extract_line_items_from_sheet
    """
    
    # Condition A: Gap Detection (Universal)
    if empty_row_count >= 10:
        logger.info(f"🛑 Stopping at row {current_row_idx}: Found {empty_row_count} consecutive empty rows")
        return True
    
    # Condition B: EPS Detection (Income Statement specific) - DEPRECATED
    # Now handled by should_stop_at_eps() before processing
    return False

def is_empty_or_formatting(text: str) -> bool:
    """
    Filter out empty or formatting-only rows
    Skip rows that are just spaces, dashes, underscores, etc.
    """
    if not text or not text.strip():
        return True
    
    # Skip rows that are just formatting characters
    cleaned = text.strip()
    if all(c in ' -_=()[]{}|\\/' for c in cleaned):
        return True
    
    # Skip very short non-meaningful text
    if len(cleaned) < 2 and not any(c.isalnum() for c in cleaned):
        return True
    
    return False

# REMOVED: extract_line_items_from_sheet function - duplicate of UniversalExcelParser._extract_line_items
# The UniversalExcelParser._extract_line_items method is the active implementation